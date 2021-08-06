from typing import Any, List
from app.zendesk_data_row import ZendeskDataRow
from pathlib import Path 
from openpyxl import load_workbook
from collections import defaultdict
from .utils import verify_path


class ZendeskDataReader:

    def __init__(self, src_file_path: str) -> None:
        # escape path 
        self._src_file_path = Path(r"{input_path}".format(input_path=src_file_path))
        self.data = None

    def load_wb(self):
        if not verify_path(self._src_file_path):
            print("Input is not a file or an excel spreadsheet.")
            exit(0)
        
        try:
            self.data = load_workbook(filename=self._src_file_path).active

        except Exception as err:
            print(f"Failed to parse with error: {err}")
            exit(0)

    # reads data from the .xlsx file at the path provided in __init__
    # and returns all rows formatted into a list/array of ZendeskDataRow
    def read_all_rows(self) -> List[ZendeskDataRow]:
        
        return [ZendeskDataRow(row) for row in self.data.iter_rows(min_row=2, values_only=True)]

# collector_container = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict, {"time": int(), "tickets":set()}))))

# for row in self.data.iter_rows(min_row=2, values_only=True):
#     '''
#     what is happening:
#     1. if no user in dict, create user key.
#     2. if no support element in dict, or no support element exists, create support element key.
#     3. if no date in dict, create date key.
#     4. create time val counter.
#     5. add time val to count, add zero if none.
#     '''

#     # collect all the data for parsing to a SAP row
#     collector_container[row[0]][(row[2] if row[2] else "N/A")][str(row[7])]["time"] += (int(row[8]) if row[8] else 0)
    
#     # add the ticket numbers on there
#     collector_container[row[0]][(row[2] if row[2] else "N/A")][str(row[7])]["tickets"].add(row[5])

# print(collector_container[list(collector_container.keys())[0]].keys())
