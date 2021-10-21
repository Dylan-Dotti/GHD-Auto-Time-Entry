from app.data_formatter.readers.zendesk.zendesk_row_factory import ZendeskRowFactory
from app.data_formatter.readers.data_row import DataRow
from openpyxl import load_workbook
from app.data_formatter.utils import date_ranges, verify_path_excel
from datetime import datetime
from pathlib import Path 
from typing import List

class ZendeskDataReader:

    def __init__(self, src_file_path: str) -> None:
        # escape path 
        self._src_file_path = self.__set_src_file_path(src_file_path)
        
        self.users = None
        self.month = None

        self.header = None
        self.data = None

    # verify the input path, raise exception if not excel
    def __set_src_file_path(self, p):
        # escape path
        t_p = Path(r"{input_path}".format(input_path=p))
        if verify_path_excel(t_p):
            return t_p
        raise AssertionError("Input is not a file or an excel spreadsheet.")

    def load_wb(self):

        try:
            self.data = load_workbook(filename=self._src_file_path).active
            self.__set_header()
            self.__set_month()
            self.__set_users()

        except Exception as err:
            raise RuntimeError(f"Failed to parse excel with error: {err}") from err

    # get the header row so we can match the fields
    def __set_header(self):
        self.header = list(next(self.data.iter_rows(min_row=1, max_row=1, values_only=True)))

    # get the users
    def get_users(self):
        return self.users

    # get the users from the spreadsheet
    def __set_users(self):
        user_col = self.header.index("Updater name")
        self.users = sorted(list(set([dt[user_col] for dt in self.data.iter_rows(min_row=2, values_only=True)])))

    # get the month
    def get_month(self):
        return self.month

    # set the month from the spreadsheet
    def __set_month(self):
        date_col = self.header.index("Update - Date")
        self.month = datetime.strptime(list(next(self.data.iter_rows(min_row=2, max_row=2, values_only=True)))[date_col], '%Y-%m-%d').date().month

    '''
      Get the available date ranges for a user.
      * a user could not have data for a week. That week shouldn't be in the week list.
    '''
    def get_weeks_with_data(self, username):
        date_col = self.header.index("Update - Date")
        user_col = self.header.index("Updater name")

        weeks = {(start, end) for start, end in date_ranges(self.get_month())}
        res = set()

        for dt in self.data.iter_rows(min_row=2, values_only=True):
             d = datetime.strptime(dt[date_col], '%Y-%m-%d').date()
             user = dt[user_col]
             if username == user: 
                for st, ed in weeks:
                    if st <= d <= ed:
                        res.add((st, ed))

        return res

    def read_all_rows(self) -> List[DataRow]:
        zendesk_row_factory = ZendeskRowFactory(self.header)
        return [zendesk_row_factory.create_row(row) for row in self.data.iter_rows(min_row=2, values_only=True)]
